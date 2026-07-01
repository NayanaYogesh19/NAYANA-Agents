import os

from openpyxl import Workbook, load_workbook

from openpyxl.styles import Font, PatternFill

from openpyxl.utils import get_column_letter


HEADERS = [

    "Category",
    "Product",
    "URL",
    "Current title",
    "Current Title Length",
    "Current Meta Description",
    "Current Descr Length",
    "What's in the Combos box",
    "Keyword",
    "Volume",
    "Intent",
    "Primary/secondary/supporting, LSI keywords",
    "Current Ranking",
    "If other Jandhyala product ranking add the URL",
    "GSC Impression",
    "GSC Clicks",
    "Suggested URL",
    "Primary KW in URL",
    "Suggested Title",
    "Title Length",
    "Suggested Meta Description",
    "Suggested Meta Description Length",
    "Primary KW in Description",
    "H1 tag with Primary KW",
    "Keywords",
    "Volume",
    "Intent",
    "H2",
    "H3-H4-H5-H6",
    "Main Image Alt Tag with Primary keyword",
    "Other images Alt tag with supporting LSI keywords",
    "Anchor text",
    "Internal Linking",
    "Choosen Supporting Blogs for Internal Linkbuilding",
    "Direct Answer for first Intro use Primary KW",
    "required Featured Snippet Answer",
    "FAQs",
    "Comments",
    "Competitor Specific URL"
]


def export_excel(results):

    # -----------------------------------
    # OUTPUT DIRECTORY
    # -----------------------------------

    output_dir = "output"

    os.makedirs(
        output_dir,
        exist_ok=True
    )

    file_path = os.path.join(

        output_dir,

        "seo_report.xlsx"
    )

    # -----------------------------------
    # LOAD EXISTING WORKBOOK
    # -----------------------------------

    if os.path.exists(file_path):

        workbook = load_workbook(file_path)

        sheet = workbook.active

        print("Existing workbook loaded")

        # -----------------------------------
        # CLEAR OLD DATA
        # KEEP HEADER ROW ONLY
        # -----------------------------------

        if sheet.max_row > 1:

            sheet.delete_rows(

                2,

                sheet.max_row
            )

            print("Old rows cleared")

    else:

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "SEO Workbook"

        print("New workbook created")

        # -----------------------------------
        # HEADER STYLE
        # -----------------------------------

        header_fill = PatternFill(

            start_color="1F4E78",

            end_color="1F4E78",

            fill_type="solid"
        )

        header_font = Font(

            bold=True,

            color="FFFFFF"
        )

        # -----------------------------------
        # ADD HEADERS
        # -----------------------------------

        for col_num, header in enumerate(HEADERS, 1):

            cell = sheet.cell(
                row=1,
                column=col_num
            )

            cell.value = header

            cell.fill = header_fill

            cell.font = header_font

    # -----------------------------------
    # START ROW
    # -----------------------------------

    row_num = sheet.max_row + 1

    # -----------------------------------
    # ADD DATA ROWS
    # -----------------------------------

    for item in results:

        # -----------------------------------
        # COLUMN A → CATEGORY
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=1
        ).value = "PRODUCT"

        # -----------------------------------
        # COLUMN B → PRODUCT NAME
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=2
        ).value = item.get(
            "product_name",
            ""
        )

        # -----------------------------------
        # COLUMN C → CLICKABLE URL
        # -----------------------------------

        url_cell = sheet.cell(
            row=row_num,
            column=3
        )

        url_cell.value = item.get(
            "url",
            ""
        )

        url_cell.hyperlink = item.get(
            "url",
            ""
        )

        url_cell.style = "Hyperlink"

        # -----------------------------------
        # COLUMN D → TITLE
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=4
        ).value = item.get(
            "title",
            ""
        )

        # -----------------------------------
        # COLUMN E → TITLE LENGTH
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=5
        ).value = item.get(
            "title_length",
            0
        )

        # -----------------------------------
        # COLUMN F → META DESCRIPTION
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=6
        ).value = item.get(
            "meta_description",
            ""
        )

        # -----------------------------------
        # COLUMN G → META LENGTH
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=7
        ).value = item.get(
            "meta_description_length",
            0
        )

        # -----------------------------------
        # COLUMN H → COMBO KEYWORDS
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=8
        ).value = ", ".join(

            item.get(
                "combo_keywords",
                []
            )
        )

        # -----------------------------------
        # COLUMN I → KEYWORD
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=9
        ).value = item.get(
            "keyword",
            ""
        )

        # -----------------------------------
        # COLUMN J → VOLUME
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=10
        ).value = item.get(
            "volume",
            0
        )

        # -----------------------------------
        # COLUMN K → INTENT
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=11
        ).value = item.get(
            "intent",
            ""
        )

        # -----------------------------------
        # COLUMN L → LSI KEYWORDS
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=12
        ).value = ", ".join(

            item.get(
                "lsi_keywords",
                []
            )
        )

        # -----------------------------------
        # COLUMN M → CURRENT RANKING
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=13
        ).value = item.get(
            "current_ranking",
            ""
        )

        # -----------------------------------
        # COLUMN Q → SUGGESTED URL
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=17
        ).value = item.get(
            "suggested_url",
            ""
        )

        # -----------------------------------
        # COLUMN S → SUGGESTED TITLE
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=19
        ).value = item.get(
            "suggested_title",
            ""
        )

        # -----------------------------------
        # COLUMN T → SUGGESTED TITLE LENGTH
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=20
        ).value = len(

            item.get(
                "suggested_title",
                ""
            )
        )

        # -----------------------------------
        # COLUMN U → SUGGESTED META
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=21
        ).value = item.get(
            "suggested_meta",
            ""
        )

        # -----------------------------------
        # COLUMN V → SUGGESTED META LENGTH
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=22
        ).value = len(

            item.get(
                "suggested_meta",
                ""
            )
        )

        # -----------------------------------
        # COLUMN X → H1
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=24
        ).value = item.get(
            "h1",
            ""
        )

        # -----------------------------------
        # SECOND KEYWORD SECTION
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=25
        ).value = item.get(
            "keyword",
            ""
        )

        sheet.cell(
            row=row_num,
            column=26
        ).value = item.get(
            "volume",
            0
        )

        sheet.cell(
            row=row_num,
            column=27
        ).value = item.get(
            "intent",
            ""
        )

        # -----------------------------------
        # COLUMN AB → H2
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=28
        ).value = ", ".join(

            item.get(
                "h2",
                []
            )
        )

        # -----------------------------------
        # COLUMN AC → H3-H6
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=29
        ).value = ", ".join(

            item.get(
                "h3_h6",
                []
            )
        )

        # -----------------------------------
        # COLUMN AD → MAIN IMAGE ALT
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=30
        ).value = item.get(
            "main_image_alt",
            ""
        )

        # -----------------------------------
        # COLUMN AE → OTHER IMAGE ALTS
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=31
        ).value = ", ".join(

            item.get(
                "other_image_alts",
                []
            )
        )

        # -----------------------------------
        # COLUMN AF → ANCHOR TEXTS
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=32
        ).value = ", ".join(

            item.get(
                "anchor_texts",
                []
            )
        )

        # -----------------------------------
        # COLUMN AG → INTERNAL LINKS
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=33
        ).value = ", ".join(

            item.get(
                "internal_links",
                []
            )
        )

        # -----------------------------------
        # COLUMN AI → FEATURED SNIPPET
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=35
        ).value = item.get(
            "featured_snippet",
            ""
        )

        # -----------------------------------
        # COLUMN AK → FAQS
        # -----------------------------------

        sheet.cell(
            row=row_num,
            column=37
        ).value = ", ".join(

            item.get(
                "faqs",
                []
            )
        )

        row_num += 1

    # -----------------------------------
    # AUTO COLUMN WIDTH
    # -----------------------------------

    for column_cells in sheet.columns:

        length = 0

        column = column_cells[0].column

        for cell in column_cells:

            try:

                if cell.value:

                    length = max(

                        length,

                        len(str(cell.value))
                    )

            except:
                pass

        adjusted_width = min(
            length + 5,
            50
        )

        sheet.column_dimensions[
            get_column_letter(column)
        ].width = adjusted_width

    # -----------------------------------
    # SAVE WORKBOOK
    # -----------------------------------

    workbook.save(file_path)

    print(f"Excel saved: {file_path}")

    return file_path